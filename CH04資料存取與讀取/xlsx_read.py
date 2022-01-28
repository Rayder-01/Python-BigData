import openpyxl
from sqlalchemy import column

# 讀取檔案
workbook = openpyxl.load_workbook('test.xlsx')

sheet = workbook.worksheets[0]

print(sheet['A1'],sheet['A1'].value)
# sheet 屬性 找出欄列數
print(sheet.max_row, sheet.max_column)

# 顯示 cell 資料
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value,end=',')
    print()

